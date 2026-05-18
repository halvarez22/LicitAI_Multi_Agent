from __future__ import annotations

import json
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from docx import Document
from openpyxl import Workbook

from app.services.document_fill_quality_gate import validate_generated_documents_fill

IssueKey = Tuple[str, str, str]


@dataclass(frozen=True)
class CalibrationCase:
    case_id: str
    stage: str
    documents: Sequence[Dict[str, Any]]
    master_profile: Dict[str, Any]
    expected_issues: Sequence[Dict[str, str]]
    expected_blocking: bool
    provenance_context: Dict[str, Any] | None = None


@dataclass(frozen=True)
class CalibrationPolicy:
    policy_version: str
    ignore_rules: Sequence[Dict[str, Any]]
    severity_overrides: Sequence[Dict[str, Any]]


def _issue_key(issue: Dict[str, Any]) -> IssueKey:
    return (
        str(issue.get("error_type") or ""),
        str(issue.get("field_key") or ""),
        str(issue.get("severity") or ""),
    )


def _materialize_docx(path: Path, lines: Sequence[str]) -> None:
    doc = Document()
    for line in lines:
        doc.add_paragraph(str(line))
    doc.save(path)


def _materialize_xlsx(path: Path, labels: Dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    row = 1
    for key in ("SUBTOTAL", "IVA", "TOTAL"):
        ws[f"A{row}"] = key
        if key in labels:
            ws[f"B{row}"] = labels[key]
        row += 1
    try:
        wb.save(path)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _materialize_case_documents(case: CalibrationCase, base_dir: Path) -> List[Dict[str, Any]]:
    generated_docs: List[Dict[str, Any]] = []
    for doc in case.documents:
        filename = str(doc.get("filename") or "").strip()
        kind = str(doc.get("kind") or "").strip().lower()
        if not filename or kind not in {"docx", "xlsx"}:
            continue
        dst = base_dir / filename
        dst.parent.mkdir(parents=True, exist_ok=True)
        if kind == "docx":
            _materialize_docx(dst, list(doc.get("lines") or []))
        else:
            _materialize_xlsx(dst, dict(doc.get("labels") or {}))
        generated_docs.append(
            {
                "ruta": str(dst),
                "tipo": doc.get("tipo"),
                "template_id": doc.get("template_id"),
            }
        )
    return generated_docs


def _confusion_from_sets(expected: Iterable[IssueKey], actual: Iterable[IssueKey]) -> Dict[str, int]:
    exp = set(expected)
    act = set(actual)
    tp = len(exp & act)
    fp = len(act - exp)
    fn = len(exp - act)
    return {"tp": tp, "fp": fp, "fn": fn}


def _safe_div(num: float, den: float) -> float:
    if den <= 0:
        return 0.0
    return num / den


def _build_metrics_by_error_type(rows: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, float]]:
    agg: Dict[str, Dict[str, int]] = {}
    for row in rows:
        for item in row.get("tp_fp_fn_by_type", []):
            key = item["error_type"]
            slot = agg.setdefault(key, {"tp": 0, "fp": 0, "fn": 0})
            slot["tp"] += int(item["tp"])
            slot["fp"] += int(item["fp"])
            slot["fn"] += int(item["fn"])

    out: Dict[str, Dict[str, float]] = {}
    for k, v in agg.items():
        tp = float(v["tp"])
        fp = float(v["fp"])
        fn = float(v["fn"])
        precision = _safe_div(tp, tp + fp)
        recall = _safe_div(tp, tp + fn)
        f1 = _safe_div(2 * precision * recall, precision + recall)
        out[k] = {
            "tp": tp,
            "fp": fp,
            "fn": fn,
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
        }
    return out


def load_calibration_cases(path: str | Path) -> List[CalibrationCase]:
    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    cases: List[CalibrationCase] = []
    for item in raw.get("cases", []):
        cases.append(
            CalibrationCase(
                case_id=str(item.get("case_id") or ""),
                stage=str(item.get("stage") or ""),
                documents=list(item.get("documents") or []),
                master_profile=dict(item.get("master_profile") or {}),
                expected_issues=list(item.get("expected_issues") or []),
                expected_blocking=bool(item.get("expected_blocking", False)),
                provenance_context=dict(item.get("provenance_context") or {}) if item.get("provenance_context") else None,
            )
        )
    return cases


def load_calibration_policy(path: str | Path) -> CalibrationPolicy:
    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    return CalibrationPolicy(
        policy_version=str(raw.get("policy_version") or "policy-unknown"),
        ignore_rules=list(raw.get("ignore_rules") or []),
        severity_overrides=list(raw.get("severity_overrides") or []),
    )


def _matches_master_profile(case: CalibrationCase, expected: Dict[str, Any]) -> bool:
    if not expected:
        return True
    mp = {str(k).lower(): str(v).lower() for k, v in (case.master_profile or {}).items()}
    for k, v in expected.items():
        if mp.get(str(k).lower()) != str(v).lower():
            return False
    return True


def _rule_matches(rule: Dict[str, Any], issue: Dict[str, Any], case: CalibrationCase) -> bool:
    if rule.get("error_type") and str(rule.get("error_type")) != str(issue.get("error_type")):
        return False
    if rule.get("field_key") and str(rule.get("field_key")) != str(issue.get("field_key")):
        return False
    if rule.get("stage") and str(rule.get("stage")) != str(case.stage):
        return False
    if rule.get("template_id"):
        template_ids = {str(d.get("template_id") or "") for d in case.documents}
        if str(rule.get("template_id")) not in template_ids:
            return False
    if not _matches_master_profile(case, dict(rule.get("when_master_profile") or {})):
        return False
    min_conf = rule.get("min_provenance_confidence")
    if min_conf is not None:
        prov = issue.get("provenance") if isinstance(issue.get("provenance"), dict) else {}
        conf = float(prov.get("confidence", 0.0) or 0.0)
        if conf < float(min_conf):
            return False
    return True


def _rule_matches_with_signals(
    rule: Dict[str, Any],
    issue: Dict[str, Any],
    case: CalibrationCase,
    signals: Dict[str, Any],
) -> bool:
    if not _rule_matches(rule, issue, case):
        return False
    expected_signals = dict(rule.get("when_signal") or {})
    for k, expected in expected_signals.items():
        if signals.get(str(k)) != expected:
            return False
    return True


def _compute_case_signals(case: CalibrationCase, actual_issues: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Calcula señales derivadas para tuning de policy basado en evidencia.
    """
    stage = str(case.stage or "").lower()
    if stage != "economic":
        return {"consistency_pass": False, "arithmetic_match": False, "arithmetic_delta": None, "tolerance": 0.02}

    def _num(v: str) -> float | None:
        txt = str(v or "").replace(",", "").strip()
        if not txt:
            return None
        try:
            return float(txt)
        except Exception:
            return None

    subtotal = None
    iva = None
    total = None
    for d in case.documents or []:
        kind = str(d.get("kind") or "").lower()
        if kind == "xlsx":
            labels = dict(d.get("labels") or {})
            subtotal = subtotal if subtotal is not None else _num(labels.get("SUBTOTAL"))
            iva = iva if iva is not None else _num(labels.get("IVA"))
            total = total if total is not None else _num(labels.get("TOTAL"))
        elif kind == "docx":
            for line in list(d.get("lines") or []):
                s = str(line or "")
                if subtotal is None:
                    m = re.search(r"^\s*SUBTOTAL\s*:\s*\$?\s*([0-9\.,]+)", s, flags=re.IGNORECASE)
                    if m:
                        subtotal = _num(m.group(1))
                if iva is None:
                    m = re.search(r"^\s*IVA(?:\s*\(.*?\))?\s*:\s*\$?\s*([0-9\.,]+)", s, flags=re.IGNORECASE)
                    if m:
                        iva = _num(m.group(1))
                if total is None:
                    m = re.search(r"^\s*TOTAL(?:\s+DE\s+LA\s+PROPUESTA)?\s*:\s*\$?\s*([0-9\.,]+)", s, flags=re.IGNORECASE)
                    if m:
                        total = _num(m.group(1))

    tolerance = 0.02
    arithmetic_delta = None
    arithmetic_match = False
    if subtotal is not None and iva is not None and total is not None:
        arithmetic_delta = abs((subtotal + iva) - total)
        arithmetic_match = arithmetic_delta <= tolerance

    has_missing_totals = any(
        str(i.get("error_type") or "") == "required_field_missing"
        and str(i.get("field_key") or "").lower() in {"subtotal", "iva", "total"}
        for i in actual_issues
    )
    consistency_pass = arithmetic_match and (not has_missing_totals)
    return {
        "consistency_pass": consistency_pass,
        "arithmetic_match": arithmetic_match,
        "arithmetic_delta": arithmetic_delta,
        "tolerance": tolerance,
    }


def _apply_policy_to_issue(
    issue: Dict[str, Any],
    case: CalibrationCase,
    policy: CalibrationPolicy | None,
    signals: Dict[str, Any],
) -> Dict[str, Any] | None:
    # Guarda blindada: placeholder en económico nunca se silencia.
    if str(case.stage or "").lower() == "economic" and str(issue.get("error_type") or "") == "placeholder_detected":
        return issue
    if policy is None:
        return issue
    for rule in policy.ignore_rules:
        if _rule_matches_with_signals(rule, issue, case, signals):
            return None
    patched = dict(issue)
    for rule in policy.severity_overrides:
        if _rule_matches_with_signals(rule, issue, case, signals):
            patched["severity"] = str(rule.get("severity") or patched.get("severity") or "")
    return patched


def _apply_policy_to_issues(
    issues: Sequence[Dict[str, Any]],
    case: CalibrationCase,
    policy: CalibrationPolicy | None,
    signals: Dict[str, Any],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for issue in issues:
        patched = _apply_policy_to_issue(issue, case, policy, signals)
        if patched is not None:
            out.append(patched)
    return out


def run_fill_gate_calibration(
    cases: Sequence[CalibrationCase],
    *,
    policy: CalibrationPolicy | None = None,
) -> Dict[str, Any]:
    rows: List[Dict[str, Any]] = []
    tmp = tempfile.mkdtemp(prefix="fill_gate_cal_")
    try:
        base = Path(tmp)
        for case in cases:
            generated_docs = _materialize_case_documents(case, base / case.case_id)
            out = validate_generated_documents_fill(
                stage=case.stage,
                generated_documents=generated_docs,
                master_profile=case.master_profile,
                provenance_context=case.provenance_context,
            )
            raw_actual_issues = list(out.get("issues", []))
            signals = _compute_case_signals(case, raw_actual_issues)
            expected_issues = _apply_policy_to_issues(list(case.expected_issues), case, policy, signals)
            actual_issues = _apply_policy_to_issues(raw_actual_issues, case, policy, signals)
            expected_keys = [_issue_key(x) for x in expected_issues]
            actual_keys = [_issue_key(x) for x in actual_issues]
            conf = _confusion_from_sets(expected_keys, actual_keys)
            expected_block = bool(case.expected_blocking)
            actual_block = any(str(i.get("severity") or "") == "block" for i in actual_issues)
            rows.append(
                {
                    "case_id": case.case_id,
                    "stage": case.stage,
                    "expected_blocking": expected_block,
                    "actual_blocking": actual_block,
                    "blocking_match": expected_block == actual_block,
                    "tp_fp_fn_by_type": [
                        {"error_type": k[0], "tp": int(k in set(expected_keys) and k in set(actual_keys)), "fp": int(k not in set(expected_keys) and k in set(actual_keys)), "fn": int(k in set(expected_keys) and k not in set(actual_keys))}
                        for k in sorted(set(expected_keys) | set(actual_keys))
                        if k[0]
                    ],
                    "tp": conf["tp"],
                    "fp": conf["fp"],
                    "fn": conf["fn"],
                    "blocking_count": sum(1 for i in actual_issues if str(i.get("severity") or "") == "block"),
                    "warning_count": sum(1 for i in actual_issues if str(i.get("severity") or "") == "warn"),
                    "signals": signals,
                }
            )
    finally:
        # En Windows algunos handlers de openpyxl pueden tardar en liberar lock.
        # Hacemos limpieza best-effort sin abortar corrida de calibración.
        shutil.rmtree(tmp, ignore_errors=True)

    tp_total = sum(r["tp"] for r in rows)
    fp_total = sum(r["fp"] for r in rows)
    fn_total = sum(r["fn"] for r in rows)
    precision = _safe_div(tp_total, tp_total + fp_total)
    recall = _safe_div(tp_total, tp_total + fn_total)
    f1 = _safe_div(2 * precision * recall, precision + recall)
    blocking_matches = sum(1 for r in rows if r["blocking_match"])
    return {
        "cases_total": len(rows),
        "policy_version": policy.policy_version if policy is not None else "calibration-v1",
        "blocking_match_rate": round(_safe_div(blocking_matches, len(rows)), 4),
        "global_metrics": {
            "tp": tp_total,
            "fp": fp_total,
            "fn": fn_total,
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
        },
        "metrics_by_error_type": _build_metrics_by_error_type(rows),
        "results": rows,
    }

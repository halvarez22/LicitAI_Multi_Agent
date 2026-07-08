import os
from typing import Any, Dict

from openpyxl import load_workbook

from app.core.logging_config import get_logger
from app.services.economic_fsr_policy import match_ingest_concept_to_fsr_key, required_fsr_param_keys

logger = get_logger(__name__)


class EconomicDataIngestor:
    """
    Ingestor y parser de plantillas Excel para captura económica universal.

    - Precios unitarios → ``economic_user_inputs.concept_prices`` (sesión).
    - Parámetros FSR canónicos → ``economic_user_inputs`` (sesión), sin perfil corporativo legacy.
    """

    @staticmethod
    def parse_economic_excel(file_path: str) -> Dict[str, Any]:
        """
        Lee el archivo Excel y extrae precios y parámetros FSR según política HRU.
        """
        results: Dict[str, Any] = {
            "fsr_params": {},
            "concept_prices": {},
            "errors": [],
        }

        if not os.path.exists(file_path):
            results["errors"].append("El archivo temporal de Excel no existe.")
            return results

        wb = None
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            header_row = 1
            concept_col = None
            value_col = None

            found_headers = False
            for r in range(1, 6):
                row_vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 10)]
                for idx, val in enumerate(row_vals, 1):
                    if val in ("concepto", "parametro", "concepto / parametro", "concept", "campo", "variable"):
                        concept_col = idx
                    if val in ("valor", "monto", "precio", "value", "price", "tasa", "costo"):
                        value_col = idx
                if concept_col and value_col:
                    header_row = r
                    found_headers = True
                    break

            if not found_headers:
                concept_col = 1
                value_col = 2
                logger.warning(
                    "[EconomicIngestor] Encabezados no detectados. Usando fallback (Col 1: Concepto, Col 2: Valor)."
                )

            max_row = ws.max_row
            for r in range(header_row + 1, max_row + 1):
                concept_val = ws.cell(row=r, column=concept_col).value
                value_val = ws.cell(row=r, column=value_col).value

                if concept_val is None:
                    continue

                concept_str = str(concept_val).strip()
                concept_lower = concept_str.lower().strip()
                fsr_key = match_ingest_concept_to_fsr_key(concept_lower)

                if fsr_key:
                    try:
                        val_num = float(str(value_val).replace(",", "").replace("$", ""))
                        results["fsr_params"][fsr_key] = val_num
                    except Exception:
                        results["errors"].append(
                            f"Fila {r}: Valor inválido para {fsr_key} ({value_val})"
                        )
                    continue

                if value_val is not None:
                    try:
                        val_num = float(str(value_val).replace(",", "").replace("$", ""))
                        results["concept_prices"][concept_str] = val_num
                    except Exception:
                        pass

            if not results["fsr_params"] and not results["concept_prices"]:
                results["errors"].append(
                    "No se pudieron extraer datos económicos del archivo. Verifica el formato."
                )

        except Exception as e:
            logger.error("[EconomicIngestor] Error procesando Excel: %s", e)
            results["errors"].append(f"Error crítico al leer el archivo Excel: {str(e)}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

        return results

    @classmethod
    async def ingest_and_save_data(
        cls, memory: Any, session_id: str, company_id: str, file_path: str
    ) -> Dict[str, Any]:
        """
        Persiste precios y parámetros FSR en ``economic_user_inputs`` de la sesión.
        """
        parse_results = cls.parse_economic_excel(file_path)

        if (
            parse_results["errors"]
            and not parse_results["fsr_params"]
            and not parse_results["concept_prices"]
        ):
            return {
                "success": False,
                "message": "Fallo al procesar el Excel.",
                "errors": parse_results["errors"],
            }

        fsr_data = parse_results.get("fsr_params") or {}
        concept_prices = parse_results.get("concept_prices") or {}
        session = await memory.get_session(session_id)
        if session and (concept_prices or fsr_data):
            inputs = session.get("economic_user_inputs", {})
            if not isinstance(inputs, dict):
                inputs = {}

            current_prices = inputs.get("concept_prices", {})
            if not isinstance(current_prices, dict):
                current_prices = {}

            for concept, price in concept_prices.items():
                current_prices[concept] = price
            inputs["concept_prices"] = current_prices

            for key in required_fsr_param_keys():
                if key in fsr_data:
                    inputs[key] = fsr_data[key]

            session["economic_user_inputs"] = inputs
            await memory.save_session(session_id, session)
            logger.info(
                "[EconomicIngestor] Precios y parámetros FSR inyectados en sesión %s",
                session_id,
            )

            from app.economic_validation.service import refresh_economic_validations_for_session

            try:
                result = await refresh_economic_validations_for_session(memory, session_id)
                if hasattr(result, "engine_output") and result.engine_output:
                    vat_amount = getattr(result.engine_output, "total_vat", "N/A")
                    subtotal = getattr(result.engine_output, "subtotal", "N/A")
                    logger.info(
                        "[EconomicIngestor] Recalculo económico OK: IVA=%s subtotal=%s",
                        vat_amount,
                        subtotal,
                    )
            except Exception as e:
                logger.error("[EconomicIngestor] Error al recalcular validaciones: %s", e)

        return {
            "success": True,
            "message": "Datos económicos importados desde la plantilla Excel.",
            "data": parse_results,
        }

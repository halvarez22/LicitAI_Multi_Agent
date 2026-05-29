
import os
import openpyxl
from typing import List, Dict, Any, Optional
from app.core.observability import get_logger

logger = get_logger(__name__)

class ExcelFillingService:
    """
    Servicio 'Espejo' para Excel: 
    Inyecta datos confirmados en archivos .xlsx originales preservando el formato oficial.
    """

    def __init__(self, base_data_dir: str = "/data"):
        self.base_data_dir = base_data_dir

    @staticmethod
    def _resolve_writable_cell(sheet: Any, row_idx: int, col_idx: int) -> tuple[int, int]:
        """Redirige a la celda ancla si el target cae dentro de un rango mergeado."""
        for merged_range in sheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col
            ):
                return merged_range.min_row, merged_range.min_col
        return row_idx, col_idx

    @staticmethod
    def _normalize_locator(row_idx: Any, col_idx: Any) -> tuple[int, int] | tuple[None, None]:
        """
        Normaliza coordenadas 0-indexed provenientes del extractor tabular.

        Devuelve ``(row, col)`` listo para openpyxl (1-indexed) o ``(None, None)``
        si el locator es inválido.
        """
        try:
            row0 = int(float(row_idx))
            col0 = int(float(col_idx))
        except (TypeError, ValueError):
            return None, None
        if row0 < 0 or col0 < 0:
            return None, None
        return row0 + 2, col0 + 1

    def fill_proposal_excel(
        self, 
        session_id: str, 
        source_filename: str, 
        items_to_fill: List[Dict[str, Any]],
        output_filename: str = None,
        source_path: str | None = None,
        output_dir: str | None = None,
    ) -> str:
        """
        Toma el archivo original de inputs e inyecta los precios en un nuevo archivo en outputs.
        
        items_to_fill debe contener dicts con:
            - sheet_name
            - row_index (0-indexed desde pandas)
            - price_column_index
            - final_price
        """
        input_path = source_path or os.path.join(
            self.base_data_dir, "inputs", session_id, source_filename
        )
        resolved_output_dir = output_dir or os.path.join(
            self.base_data_dir, "outputs", session_id, "2.propuesta_economica"
        )
        os.makedirs(resolved_output_dir, exist_ok=True)
        
        if not output_filename:
            output_filename = f"PROPUESTA_ECONOMICA_{source_filename}"
        
        output_path = os.path.join(resolved_output_dir, output_filename)

        if not os.path.exists(input_path):
            logger.error("excel_fill_input_not_found", path=input_path)
            raise FileNotFoundError(f"No se encontró el archivo original: {source_filename}")

        try:
            # Cargar el libro original preservando estilos
            wb = openpyxl.load_workbook(input_path, data_only=False)
            
            filled_count = 0
            sheet_subtotals: Dict[str, float] = {}

            for item in items_to_fill:
                sheet_name = item.get("sheet_name")
                row_idx = item.get("row_index")
                col_idx = item.get("price_column_index")
                price = item.get("final_price")

                if sheet_name not in wb.sheetnames:
                    logger.warning("excel_fill_sheet_not_found", sheet=sheet_name)
                    continue
                
                sheet = wb[sheet_name]
                
                # Ajuste de índices: 
                # pandas es 0-indexed y excluye el header si lo detectó.
                # openpyxl es 1-indexed.
                # Si pandas detectó headers, la fila 0 de pandas es la fila 2 de Excel.
                # Por seguridad, asumimos que row_index ya viene ajustado o lo ajustamos aquí + 2.
                # NOTA: En tabular_line_item_extract, el row_index es el índice del DataFrame.
                excel_row, excel_col = self._normalize_locator(row_idx, col_idx)
                if excel_row is None or excel_col is None:
                    logger.warning(
                        "excel_fill_invalid_locator",
                        sheet=sheet_name,
                        row_index=row_idx,
                        price_column_index=col_idx,
                    )
                    continue
                
                try:
                    excel_row, excel_col = self._resolve_writable_cell(sheet, excel_row, excel_col)
                    sheet.cell(row=excel_row, column=excel_col).value = price
                    filled_count += 1

                    qty = item.get("quantity")
                    amt_col = item.get("amount_column_index")
                    if qty is not None and amt_col is not None:
                        try:
                            amount = float(price) * float(qty)
                            amt_excel_row, amt_excel_col = self._normalize_locator(
                                row_idx, amt_col
                            )
                            if amt_excel_row is not None and amt_excel_col is not None:
                                amt_excel_row, amt_excel_col = self._resolve_writable_cell(
                                    sheet, amt_excel_row, amt_excel_col
                                )
                                sheet.cell(row=amt_excel_row, column=amt_excel_col).value = round(
                                    amount, 2
                                )
                                sheet_subtotals[sheet_name] = (
                                    sheet_subtotals.get(sheet_name, 0.0) + amount
                                )
                        except (TypeError, ValueError):
                            pass
                    elif price is not None:
                        try:
                            sheet_subtotals[sheet_name] = sheet_subtotals.get(
                                sheet_name, 0.0
                            ) + float(price)
                        except (TypeError, ValueError):
                            pass
                except Exception as e:
                    logger.error("excel_fill_cell_error", row=excel_row, col=excel_col, error=str(e))

            self._write_sheet_totals_if_needed(wb, items_to_fill, sheet_subtotals)

            wb.save(output_path)
            logger.info("excel_fill_completed", session_id=session_id, filled_items=filled_count, output=output_path)
            return output_path

        except Exception as e:
            logger.error("excel_fill_critical_error", session_id=session_id, error=str(e))
            raise e

    @staticmethod
    def _write_sheet_totals_if_needed(
        wb: Any,
        items_to_fill: List[Dict[str, Any]],
        sheet_subtotals: Dict[str, float],
    ) -> None:
        """
        Escribe total de hoja en celda de total detectada cuando la plantilla no trae fórmulas.
        """
        total_targets: Dict[str, tuple] = {}
        for item in items_to_fill or []:
            sname = item.get("sheet_name")
            tcol = item.get("total_column_index")
            if sname and tcol is not None and sname not in total_targets:
                total_targets[str(sname)] = (item.get("row_index"), tcol)

        for sheet_name, subtotal in (sheet_subtotals or {}).items():
            if sheet_name not in wb.sheetnames or subtotal <= 0:
                continue
            target = total_targets.get(sheet_name)
            if not target:
                continue
            row_idx, col_idx = target
            excel_row, excel_col = ExcelFillingService._normalize_locator(row_idx, col_idx)
            if excel_row is None:
                continue
            sheet = wb[sheet_name]
            scan_start = max(1, excel_row)
            for r in range(scan_start, min(scan_start + 40, sheet.max_row + 1)):
                excel_row, excel_col = ExcelFillingService._resolve_writable_cell(
                    sheet, r, excel_col
                )
                sheet.cell(row=excel_row, column=excel_col).value = round(subtotal, 2)
                break

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.excel_filling_service import ExcelFillingService


def test_fill_proposal_excel_redirects_merged_cells_to_anchor(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.merge_cells("H9:I9")
    ws["H9"] = "placeholder"
    wb.save(source)

    service = ExcelFillingService(base_data_dir=str(tmp_path))
    output = service.fill_proposal_excel(
        session_id="sess",
        source_filename="source.xlsx",
        source_path=str(source),
        output_dir=str(tmp_path / "out"),
        items_to_fill=[
            {
                "sheet_name": "Hoja1",
                "row_index": 7,  # -> fila Excel 9
                "price_column_index": 8,  # -> columna Excel 9 (I), dentro del merge H9:I9
                "final_price": 123.45,
            }
        ],
    )

    out_wb = load_workbook(output, data_only=False)
    out_ws = out_wb["Hoja1"]
    assert out_ws["H9"].value == 123.45
    out_wb.close()
from pathlib import Path

import openpyxl

from app.services.excel_filling_service import ExcelFillingService


def _make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partidas"
    ws.cell(row=1, column=1, value="Concepto")
    ws.cell(row=1, column=2, value="Precio")
    ws.cell(row=2, column=1, value="Servicio A")
    ws.cell(row=3, column=1, value="Servicio B")
    wb.save(path)


def test_fill_proposal_excel_escribe_precio_en_coordenada_valida(tmp_path: Path) -> None:
    source = tmp_path / "catalogo.xlsx"
    _make_workbook(source)
    svc = ExcelFillingService(base_data_dir=str(tmp_path))

    output = svc.fill_proposal_excel(
        session_id="s1",
        source_filename="catalogo.xlsx",
        source_path=str(source),
        output_filename="salida.xlsx",
        items_to_fill=[
            {
                "sheet_name": "Partidas",
                "row_index": 0,
                "price_column_index": 1,
                "final_price": 123.45,
            }
        ],
    )

    wb = openpyxl.load_workbook(output, data_only=False)
    assert "2.propuesta_economica" in output
    assert wb["Partidas"].cell(row=2, column=2).value == 123.45


def test_fill_proposal_excel_omite_locator_invalido(tmp_path: Path) -> None:
    source = tmp_path / "catalogo.xlsx"
    _make_workbook(source)
    svc = ExcelFillingService(base_data_dir=str(tmp_path))

    output = svc.fill_proposal_excel(
        session_id="s2",
        source_filename="catalogo.xlsx",
        source_path=str(source),
        output_filename="salida_invalida.xlsx",
        items_to_fill=[
            {
                "sheet_name": "Partidas",
                "row_index": 0,
                "price_column_index": -1,
                "final_price": 999.0,
            }
        ],
    )

    wb = openpyxl.load_workbook(output, data_only=False)
    assert wb["Partidas"].cell(row=2, column=1).value == "Servicio A"
    assert wb["Partidas"].cell(row=2, column=2).value is None


def test_fill_proposal_excel_writes_amount_and_sheet_total(tmp_path: Path) -> None:
    source = tmp_path / "grid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ZB"
    ws.cell(row=1, column=1, value="Localidad")
    ws.cell(row=1, column=2, value="Cantidad")
    ws.cell(row=1, column=3, value="Precio")
    ws.cell(row=1, column=4, value="Importe")
    ws.cell(row=1, column=5, value="Total")
    ws.cell(row=2, column=1, value="Acámbaro")
    ws.cell(row=2, column=2, value=2)
    ws.cell(row=10, column=5, value="")
    wb.save(source)

    svc = ExcelFillingService(base_data_dir=str(tmp_path))
    output = svc.fill_proposal_excel(
        session_id="s3",
        source_filename="grid.xlsx",
        source_path=str(source),
        output_filename="grid_out.xlsx",
        items_to_fill=[
            {
                "sheet_name": "ZB",
                "row_index": 0,
                "price_column_index": 2,
                "quantity": 2,
                "amount_column_index": 3,
                "total_column_index": 4,
                "final_price": 100.0,
            }
        ],
    )

    out = openpyxl.load_workbook(output, data_only=False)
    sheet = out["ZB"]
    assert sheet.cell(row=2, column=3).value == 100.0
    assert sheet.cell(row=2, column=4).value == 200.0
    assert sheet.cell(row=2, column=5).value == 200.0

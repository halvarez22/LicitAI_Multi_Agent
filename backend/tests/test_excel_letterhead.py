"""Membrete corporativo en Excel de propuesta económica."""
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.utils.doc_formatting import apply_corporate_excel_letterhead


def test_apply_corporate_excel_letterhead_inserts_image(tmp_path):
    logo = tmp_path / "logo.png"
    logo.write_bytes(
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
        b"\x01\x01\x01\x00\x18\xdd\x8d\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    wb = Workbook()
    ws = wb.active
    header_row = apply_corporate_excel_letterhead(
        ws,
        {
            "logo_path": str(logo),
            "empresa": "Empresa Test SA",
            "rfc": "TST010101TST",
            "tender_name": "LICITACION DEMO",
            "fecha_corta": "23/04/2026",
        },
    )
    assert header_row == 5
    ws.cell(row=header_row, column=1, value="Partida")
    out = tmp_path / "tabla.xlsx"
    wb.save(out)

    wb2 = load_workbook(out)
    ws2 = wb2.active
    assert len(ws2._images) >= 1
    assert "EMPRESA TEST" in str(ws2.cell(row=1, column=2).value or "").upper()

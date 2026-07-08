from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import Workbook

from app.config.settings import settings as app_settings
from app.services.document_fill_quality_gate import validate_generated_documents_fill


def _make_docx(path: Path, lines: list[str]) -> None:
    doc = Document()
    for line in lines:
        doc.add_paragraph(line)
    doc.save(path)


def _make_xlsx(path: Path, with_values: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "SUBTOTAL"
    ws["A2"] = "IVA"
    ws["A3"] = "TOTAL"
    if with_values:
        ws["B1"] = 100.0
        ws["B2"] = 16.0
        ws["B3"] = 116.0
    wb.save(path)


def test_fill_gate_apu_ellipsis_en_formats_no_bloquea(tmp_path, monkeypatch):
    """Celdas '...' del análisis de precios se completan en etapa económica."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "analisis_precios_unitarios.docx"
    _make_docx(f, ["...", "Partida 1", "..."])
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f)}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    assert out["validation_passed"] is True
    assert out["blocking_count"] == 0


def test_fill_gate_tarifa_mensual_en_formats_no_bloquea(tmp_path, monkeypatch):
    """Anexo D-III: tarifa se completa en propuesta económica; no frena formatos."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    monkeypatch.setattr(app_settings, "ADMIN_ECONOMIC_DEFERRAL", True)
    f = tmp_path / "anexo_d_iii.docx"
    _make_docx(f, ["Tarifa mensual para horario: _______________"])
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f)}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    assert out["validation_passed"] is True
    assert out["blocking_count"] == 0
    assert any(
        i.get("field_key") == "tarifa_mensual" and i.get("severity") == "warn"
        for i in out["issues"]
    )


def test_fill_gate_detecta_placeholder_docx(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "x.docx"
    _make_docx(f, ["Representante: [NOMBRE]"])
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f)}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    assert out["documents_scanned"] == 1
    assert any(i["error_type"] == "placeholder_detected" for i in out["issues"])


def test_fill_gate_xlsx_anchor_labels(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "econ.xlsx"
    _make_xlsx(f, with_values=False)
    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[{"ruta": str(f), "tipo": "tabla_precios", "template_id": "tabla_precios"}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    assert out["documents_scanned"] == 1
    assert any(i["error_type"] == "required_field_missing" for i in out["issues"])


def test_fill_gate_policy_registry_y_version(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "ANEXO_AE_PROPUESTA_ECONOMICA.docx"
    _make_docx(
        f,
        [
            "ACME SA DE CV",
            "RFC: ACM010101AAA",
            "SUBTOTAL: $100.00",
            "I.V.A. (16%): $16.00",
            "TOTAL DE LA PROPUESTA: $116.00",
        ],
    )
    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[
            {
                "ruta": str(f),
                "tipo": "anexo_economico",
                "template_id": "anexo_economico",
                "materialization_route": "deterministic",
            }
        ],
        master_profile={"razon_social": "ACME SA DE CV", "rfc": "ACM010101AAA", "representante_legal": "Ana Perez"},
        provenance_context={"economic_resumen": {"subtotal": 100, "iva": 16, "total": 116}},
    )
    assert out["policy_version"] == "1.4.0"
    assert out["documents_with_policy"] == 1
    blocks = [i for i in out.get("issues") or [] if i.get("severity") == "block"]
    assert out["validation_passed"] is True, blocks


def test_fill_gate_economic_xlsx_totals_with_colon_labels(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "TABLA_PRECIOS_UNITARIOS.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["E10"] = "SUBTOTAL:"
    ws["F10"] = 100.0
    ws["E11"] = "IVA (16%):"
    ws["F11"] = 16.0
    ws["E12"] = "TOTAL:"
    ws["F12"] = 116.0
    wb.save(f)
    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[
            {
                "ruta": str(f),
                "tipo": "tabla_precios",
                "template_id": "tabla_precios",
                "materialization_route": "deterministic",
            }
        ],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
        provenance_context={"economic_resumen": {"subtotal": 100, "iva": 16, "total": 116}},
    )
    assert out["validation_passed"] is True
    assert out["blocking_count"] == 0


def test_fill_gate_carta_compromiso_trusts_profile_on_deterministic_route(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "CARTA_COMPROMISO_PRECIOS.docx"
    _make_docx(
        f,
        [
            "Comercializadora Mayo y Torres SA de CV",
            "Representante Legal: Juan Pérez",
        ],
    )
    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[
            {
                "ruta": str(f),
                "tipo": "carta_compromiso",
                "template_id": "carta_compromiso",
                "materialization_route": "deterministic",
            }
        ],
        master_profile={
            "razon_social": "Comercializadora Mayo y Torres SA de CV",
            "rfc": "CMT160107S83",
            "representante_legal": "Juan Pérez",
        },
    )
    assert out["validation_passed"] is True


def test_fill_gate_propuesta_tecnica_con_acentos_no_bloquea_por_confianza(tmp_path, monkeypatch):
    """RFC/razón social en DOCX con acentos deben reconocerse sin falso bloqueo."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "02_TE-03_Propuesta_tecnica_describiendo_especificaciones.docx"
    _make_docx(
        f,
        [
            "SEGURIDAD PRIVADA INTEGRAL MANAVIL SOCIEDAD ANÓNIMA DE CAPITAL VARIABLE",
            "RFC: SPI060200AG5",
            "Representante: YUNUEN IVON ACEVES SANCHEZ",
            "Se anexan contratos probatorios en el expediente de la empresa.",
        ],
    )
    out = validate_generated_documents_fill(
        stage="technical",
        generated_documents=[{"ruta": str(f), "tipo": "tecnico", "nombre": "TE-03"}],
        master_profile={
            "razon_social": "SEGURIDAD PRIVADA INTEGRAL MANAVIL SOCIEDAD ANONIMA DE CAPITAL VARIABLE",
            "rfc": "SPI060200AG5",
            "representante_legal": "YUNUEN IVON ACEVES SANCHEZ",
        },
        provenance_context={"source": "technical_writer", "confidence": 0.7},
    )
    error_types = {i["error_type"] for i in out["issues"]}
    assert "source_confidence_insufficient" not in error_types
    missing_fields = {i["field_key"] for i in out["issues"] if i["error_type"] == "required_field_missing"}
    assert "rfc" not in missing_fields
    assert "razon_social" not in missing_fields


def test_fill_gate_confianza_insuficiente_y_pattern(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "anexo_7.docx"
    _make_docx(f, ["ACME SA DE CV", "RFC_MALO", "Ana"])
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f), "tipo": "administrativo", "template_id": "anexo_7"}],
        master_profile={"razon_social": "ACME SA DE CV", "rfc": "RFC_MALO", "representante_legal": "Ana"},
        provenance_context={
            "source": "llm_inference",
            "field_provenance": {
                "rfc": {"source": "llm_inference", "confidence": 0.2, "anchor": None},
            },
        },
    )
    error_types = {i["error_type"] for i in out["issues"]}
    assert "cross_field_inconsistency" in error_types


def test_fill_gate_valida_que_el_docx_materializado_contenga_rfc(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "anexo_7.docx"
    _make_docx(f, ["Documento legal sin datos de empresa"])
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f), "tipo": "administrativo", "template_id": "anexo_7"}],
        master_profile={
            "razon_social": "ACME SA DE CV",
            "rfc": "ACM010101AAA",
            "representante_legal": "Ana Perez",
        },
    )
    missing_fields = {i["field_key"] for i in out["issues"] if i["error_type"] == "required_field_missing"}
    assert "rfc" in missing_fields
    assert "razon_social" in missing_fields
    assert "representante_legal" in missing_fields


def test_fill_gate_ignora_separador_pero_detecta_blanco_real(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "blancos.docx"
    _make_docx(
        f,
        [
            "__________________________________________________",
            "Quien suscribe _____________________, bajo protesta de decir verdad manifiesto lo siguiente:",
        ],
    )
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f)}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    placeholders = [i for i in out["issues"] if i["error_type"] == "placeholder_detected"]
    assert len(placeholders) == 1
    assert "Quien suscribe" in placeholders[0]["detected_value"]


def test_fill_gate_detecta_cross_tender_y_blancos_en_xlsx(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "econ_cross.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = (
        "SERVICIO INTEGRAL ... PARA LOS SERVICIOS DE SALUD DEL INSTITUTO MEXICANO "
        "DEL SEGURO SOCIAL PARA EL BIENESTAR (IMSS-BIENESTAR)."
    )
    ws["A2"] = (
        "IMSS-BIENESTAR únicamente aceptará cubrir lo manifestado en la propuesta económica."
    )
    ws["A3"] = (
        "Quien suscribe, _______________________, en mi carácter de representante legal "
        "de la empresa ______________________"
    )
    wb.save(f)
    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[{"ruta": str(f)}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
        provenance_context={"source": "economic_writer", "confidence": 0.95, "session_hint": "isapeg ISAPEG"},
    )
    error_types = {i["error_type"] for i in out["issues"]}
    assert "placeholder_detected" in error_types
    assert "cross_tender_reference" in error_types


def test_fill_gate_no_confunde_ellipsis_en_domicilio_xlsx(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "econ_dir.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "DOMICILIO"
    ws["A2"] = (
        "CARRETERA DOLORES HIDALGO-XOCONOXTLE.750...SAN ANTONIO DEL PRETORIO."
        "DOLORES HIDALGO CUNA DE LA INDEPENDENCIA NACIONAL"
    )
    wb.save(f)

    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[{"ruta": str(f), "tipo": "plantilla_economica_espejo"}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
        provenance_context={"source": "economic_writer", "confidence": 0.95, "session_hint": "isapeg ISAPEG"},
    )

    assert not any(i["error_type"] == "placeholder_detected" for i in out["issues"])


def test_fill_gate_no_aplica_totales_a_excel_espejo_solo_por_nombre(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "ECON_09_anexo_iii_p1_2_za_propuesta_economica.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ANEXO III PARTIDA 2 ZONA A PROPUESTA ECONOMICA"
    ws["B2"] = "Descripción del material"
    ws["E2"] = "Cantidad mensual"
    wb.save(f)

    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[{"ruta": str(f), "tipo": "plantilla_economica_espejo"}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
        provenance_context={"source": "economic_writer", "confidence": 0.95, "session_hint": "isapeg ISAPEG"},
    )

    required_missing = {
        i["field_key"]
        for i in out["issues"]
        if i["error_type"] == "required_field_missing"
    }
    assert "subtotal" not in required_missing
    assert "iva" not in required_missing
    assert "total" not in required_missing


def test_fill_gate_bloquea_excel_espejo_sin_locators_validos(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "econ_skipped.xlsx"
    wb = Workbook()
    wb.save(f)

    out = validate_generated_documents_fill(
        stage="economic",
        generated_documents=[
            {
                "ruta": str(f),
                "tipo": "plantilla_economica_espejo",
                "fill_status": "skipped_missing_locator",
                "valid_locator_count": 0,
            }
        ],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
        provenance_context={"source": "economic_writer", "confidence": 0.95, "session_hint": "isapeg ISAPEG"},
    )

    missing_fields = {
        i["field_key"]
        for i in out["issues"]
        if i["error_type"] == "required_field_missing"
    }
    assert "price_fill" in missing_fields


def test_fill_gate_ignora_placeholders_de_formato_referencia_fianza(tmp_path, monkeypatch):
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "anexo_g_fianza.docx"
    _make_docx(
        f,
        [
            "AUTORIZACIÓN DEL GOBIERNO FEDERAL PARA OPERAR: ________ (NÚMERO DE OFICIO Y FECHA)",
            "NÚMERO: ____________________. (NÚMERO ASIGNADO POR LA \"AFIANZADORA\")",
            "MONTO AFIANZADO: ____________ (CON LETRA Y NÚMERO, SIN INCLUIR EL IMPUESTO AL VALOR AGREGADO)",
            "FECHA DE EXPEDICIÓN: _____________________.",
            "LA \"AFIANZADORA\" garantiza el cumplimiento del contrato.",
        ],
    )
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f), "tipo": "administrativo"}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    assert not any(i["error_type"] == "placeholder_detected" for i in out["issues"])


def test_fill_gate_anexo9_titulo_subrayado_no_bloquea(tmp_path, monkeypatch):
    """Subrayado decorativo en título + IVA parcial se difiere a etapa económica."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "Anexo_9_A_RESUMEN_DE_LA_COTIZACION_MENSUAL_y_ANUAL_SUBTOTAL.docx"
    _make_docx(
        f,
        [
            "ANEXO 9: A RESUMEN DE LA COTIZACIÓN MENSUAL Y ANUAL SUBTOTAL__________________________________________ 16%…",
            "SEGURIDAD PRIVADA INTEGRAL MANAVIL SA DE CV",
            "RFC: SPI060200AG5",
        ],
    )
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f), "nombre": f.name}],
        master_profile={
            "razon_social": "SEGURIDAD PRIVADA INTEGRAL MANAVIL SA DE CV",
            "rfc": "SPI060200AG5",
            "representante_legal": "YUNUEN IVON ACEVES SANCHEZ",
        },
    )
    assert out["validation_passed"] is True
    assert out["blocking_count"] == 0


def test_fill_gate_contrato_con_fuga_prompt_no_bloquea_por_na(tmp_path, monkeypatch):
    """Restos de REGLA CRÍTICA en contrato no deben disparar falso positivo N/A."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "Modelo_de_Contrato_Abierto_según_anexo_No_14.docx"
    _make_docx(
        f,
        [
            "DECLARACIÓN BAJA PROTESTA DE DECIR VERDAD Nosotros, ACME SA DE CV, RFC ACM010101AAA. "
            "(TEXTO ESTRICTO) REGLA CRÍTICA Si no tienes un dato real verificado en el contexto, "
            'NO escribas "...", "N/A", "", "" ni placeholders entre corchetes.',
            "YUNUEN IVON ACEVES SANCHEZ\nREPRESENTANTE LEGAL\nACME SA DE CV\nR.F.C. ACM010101AAA",
        ],
    )
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f), "nombre": f.name}],
        master_profile={
            "razon_social": "ACME SA DE CV",
            "rfc": "ACM010101AAA",
            "representante_legal": "YUNUEN IVON ACEVES SANCHEZ",
        },
    )
    assert out["validation_passed"] is True
    assert out["blocking_count"] == 0


def test_cross_tender_ignores_pliego_boilerplate_materiales_anexo_dolares(tmp_path, monkeypatch):
    """MATERIALES, ANEXO y DOLARES en paréntesis son jerga de formato, no otra licitación."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    ctx = {
        "source": "formats",
        "confidence": 0.95,
        "session_hint": "barda_primaria_lopez_rayon D/080/2025",
    }
    for name, lines in (
        (
            "Anexo_E-5_Materiales.docx",
            ["Anexo E-5 (MATERIALES)", "Detalle (MATERIALES)", "Empresa SA", "RFC ACM010101AAA"],
        ),
        (
            "Aviso_de_privacidad_anexo.docx",
            ["Aviso (ANEXO)", "Referencia (ANEXO)", "Empresa SA", "RFC ACM010101AAA"],
        ),
        (
            "Capital_Contable_comprometido.docx",
            ["Capital (DOLARES)", "Importe (DOLARES)", "Empresa SA", "RFC ACM010101AAA"],
        ),
    ):
        f = tmp_path / name
        _make_docx(f, lines)
        out = validate_generated_documents_fill(
            stage="formats",
            generated_documents=[{"ruta": str(f), "nombre": f.name}],
            master_profile={"razon_social": "Empresa SA", "rfc": "ACM010101AAA"},
            provenance_context=ctx,
        )
        cross = [
            i
            for i in out["issues"]
            if i.get("error_type") == "cross_tender_reference"
        ]
        assert cross == [], f"{name} no debe marcar cross_tender: {cross}"


def test_cross_tender_still_flags_real_zone_acronym_cinco(tmp_path, monkeypatch):
    """Siglas de zona/institución (CINCO) siguen bloqueando si no están en la sesión."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    f = tmp_path / "propuesta_servicio_hospital.docx"
    _make_docx(
        f,
        [
            "Proyecto (CINCO) en Guanajuato",
            "Atención en (CINCO)",
            "Empresa SA",
            "RFC ACM010101AAA",
        ],
    )
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f), "nombre": f.name}],
        master_profile={"razon_social": "Empresa SA", "rfc": "ACM010101AAA"},
        provenance_context={
            "source": "formats",
            "confidence": 0.95,
            "session_hint": "barda_primaria_lopez_rayon D/080/2025",
        },
    )
    assert any(
        i.get("error_type") == "cross_tender_reference" for i in out["issues"]
    )


def test_fill_gate_price_fill_xlsx_en_formats_deferido(tmp_path, monkeypatch):
    """Excel de cálculo sin locators: warning en formats con ADMIN_ECONOMIC_DEFERRAL."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    monkeypatch.setattr(app_settings, "ADMIN_ECONOMIC_DEFERRAL", True)
    f = tmp_path / "calculo_costos.xlsx"
    _make_xlsx(f, with_values=False)
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[
            {
                "ruta": str(f),
                "fill_status": "skipped_missing_locator",
                "valid_locator_count": 0,
            }
        ],
        master_profile={
            "razon_social": "Comercializadora Mayo y Torres",
            "rfc": "CMT160107S83",
            "representante_legal": "Ana Torres",
        },
    )
    assert out["validation_passed"] is True
    assert out["blocking_count"] == 0
    deferred = [
        i
        for i in out["issues"]
        if i.get("expected_rule") == "deferred_to_economic_stage"
    ]
    assert deferred


def test_fill_gate_tarifa_mensual_formats_bloquea_sin_deferral(tmp_path, monkeypatch):
    """Con ADMIN_ECONOMIC_DEFERRAL=false, placeholder de tarifa sí bloquea formats."""
    monkeypatch.setattr(app_settings, "DOCUMENT_FILL_QUALITY_GATE_MODE", "enforce")
    monkeypatch.setattr(app_settings, "ADMIN_ECONOMIC_DEFERRAL", False)
    f = tmp_path / "anexo_d_iii.docx"
    _make_docx(f, ["Tarifa mensual para horario: _______________"])
    out = validate_generated_documents_fill(
        stage="formats",
        generated_documents=[{"ruta": str(f)}],
        master_profile={"razon_social": "ACME", "rfc": "ACM010101AAA", "representante_legal": "Ana"},
    )
    assert out["validation_passed"] is False
    assert out["blocking_count"] >= 1
